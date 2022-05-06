import os
import openpyxl
from openpyxl.styles import *


from web_keys.keys import Key

# 获取单元格信息，处理参数
def get_data(value):
    if values[2] is not None:
        data = {}
        # 字符串切割
        str_temp = value.split(';')
        for temp in str_temp:
            # 切割1次
            temp = temp.split('=', 1)
            data[temp[0]] = temp[1]
    else:
        data = None
    return data


data_path = "../data/"
for f in os.listdir(data_path):
    if f.startswith('~'):
        pass
    else:
        domain = os.path.abspath(data_path)
        path = os.path.join(domain, f)

        wb = openpyxl.load_workbook(path)
        # sheetname = wb['Sheet1']
        for name in wb.sheetnames:
            sheet = wb[name]
            max_row = sheet.max_row

            print("*********************{}*********************".format(name))
            for values in sheet.values:
                rows = max_row - 1
                if type(values[0]) is int:
                    print("正在执行操作步骤{}:  {}".format(values[1], values[3]))

                    # 操作步骤关联参数
                    data = get_data(values[2])
                    print(data)
                    print(sheet.max_row)


                    # （优化后弃用）当excel为多参数时，判断为none的参数清理
                    # 因为遍历中不能修改字典元素，所以将字典改为list（）
                    # for k in list(data.keys()):
                    #     if data[k] is None:
                    #         del data[k]

                    # 管理操作行为,传参
                    # 1.实例化key 对象
                    if values[1] == 'open_browser':
                        key = Key(**data)
                    elif 'assert' in values[1]:
                        status = getattr(key, values[1])(**data)
                        if status:
                            sheet.cell(row=rows, column=5).value = 'PASS'
                            sheet.cell.fill = PatternFill(fill_type='solid', fgColor="FF00FF00")
                            wb.save(path)
                        else:
                            cel = sheet.cell(row=rows, column=5)
                            cel.fill = PatternFill(fill_type='solid', fgColor="FFFF0000")
                            cel.value = 'FAIL'
                            wb.save(path)

                    else:
                        if data is not None:
                            getattr(key, values[1])(**data)
                        else:
                            getattr(key, values[1])()


